import openpyxl
from datetime import datetime
from typing import Optional
from models import InvoiceAllocation, load_structure
from config import TEMPLATE_PATH
from database import save_invoice as db_save_invoice


def create_allocations(
    supplier: str,
    invoice_template: str,
    invoice_number: str,
    invoice_date: str,
    invoice_value: float,
    drive_link: str,
    distributions: list[dict]  # [{company, brand, department, subdepartment, allocation}]
) -> list[InvoiceAllocation]:
    """
    Create allocation records for an invoice distributed across departments.

    distributions: list of dicts with keys:
        - company: str
        - brand: str (optional)
        - department: str
        - subdepartment: str (optional)
        - allocation: float (percentage as decimal, e.g., 0.5 for 50%)
    """
    # Validate allocations sum to 1 (100%)
    total_allocation = sum(d['allocation'] for d in distributions)
    if abs(total_allocation - 1.0) > 0.001:
        raise ValueError(f"Allocations must sum to 100%, got {total_allocation * 100}%")

    structure = load_structure()
    allocations = []
    submission_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for dist in distributions:
        # Find the responsible person from structure
        responsible = ''
        for unit in structure:
            if (unit.company == dist['company'] and
                unit.department == dist['department'] and
                (dist.get('subdepartment') is None or unit.subdepartment == dist.get('subdepartment'))):
                responsible = unit.manager
                break

        allocation = InvoiceAllocation(
            submission_date=submission_date,
            company=dist['company'],
            supplier=supplier,
            invoice_template=invoice_template,
            invoice_number=invoice_number,
            invoice_date=invoice_date,
            invoice_value=invoice_value,
            allocation=dist['allocation'],
            department=dist['department'],
            subdepartment=dist.get('subdepartment'),
            brand=dist.get('brand'),
            responsible=responsible,
            drive_link=drive_link,
            reinvoice_to=dist.get('reinvoice_to')
        )
        allocations.append(allocation)

    return allocations


def save_invoice_to_db(
    supplier: str,
    invoice_template: str,
    invoice_number: str,
    invoice_date: str,
    invoice_value: float,
    currency: str,
    drive_link: str,
    distributions: list[dict]
) -> int:
    """Save invoice and allocations to database. Returns invoice ID."""
    # Validate allocations sum to 1 (100%)
    total_allocation = sum(d['allocation'] for d in distributions)
    if abs(total_allocation - 1.0) > 0.001:
        raise ValueError(f"Allocations must sum to 100%, got {total_allocation * 100}%")

    # Add responsible person to each distribution
    structure = load_structure()
    for dist in distributions:
        for unit in structure:
            if (unit.company == dist['company'] and
                unit.department == dist['department'] and
                (dist.get('subdepartment') is None or unit.subdepartment == dist.get('subdepartment'))):
                dist['responsible'] = unit.manager
                break

    return db_save_invoice(
        supplier=supplier,
        invoice_template=invoice_template,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        invoice_value=invoice_value,
        currency=currency,
        drive_link=drive_link,
        distributions=distributions
    )


def export_to_template(allocations: list[InvoiceAllocation], output_path: Optional[str] = None) -> str:
    """
    Export allocations to the Template Excel file's Data sheet.
    Returns the path to the output file.
    """
    if output_path is None:
        output_path = TEMPLATE_PATH

    wb = openpyxl.load_workbook(output_path)
    ws = wb['Data']

    # Find the next empty row
    next_row = ws.max_row + 1

    for alloc in allocations:
        ws.cell(row=next_row, column=1, value=alloc.submission_date)
        ws.cell(row=next_row, column=2, value=alloc.company)
        ws.cell(row=next_row, column=3, value=alloc.supplier)
        ws.cell(row=next_row, column=4, value=alloc.invoice_template)
        ws.cell(row=next_row, column=5, value=alloc.invoice_number)
        ws.cell(row=next_row, column=6, value=alloc.invoice_date)
        ws.cell(row=next_row, column=7, value=alloc.invoice_value)
        ws.cell(row=next_row, column=8, value=alloc.allocation)
        ws.cell(row=next_row, column=9, value=alloc.department)
        ws.cell(row=next_row, column=10, value=alloc.subdepartment or '')
        ws.cell(row=next_row, column=11, value=alloc.brand or '')
        ws.cell(row=next_row, column=12, value=alloc.responsible)
        ws.cell(row=next_row, column=13, value=alloc.drive_link)
        ws.cell(row=next_row, column=14, value=alloc.reinvoice_to or '')
        next_row += 1

    wb.save(output_path)
    wb.close()
    return output_path


def get_existing_data() -> list[dict]:
    """Read existing data from the Template file."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True)
    ws = wb['Data']

    headers = []
    data = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h else f'col_{j}' for j, h in enumerate(row)]
        else:
            if any(cell is not None for cell in row):
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = cell
                data.append(row_dict)

    wb.close()
    return data


def normalize_vat(vat: str) -> str:
    """
    Normalize VAT number for comparison.
    Handles various formats: 'RO 225615', 'RO225615', 'CUI 225615', '225615', etc.
    Returns just the numeric part for comparison, plus stores the country code.
    """
    import re

    if not vat:
        return ''

    # Convert to uppercase and strip
    vat = str(vat).upper().strip()

    # Remove common prefixes like 'CUI', 'CIF', 'VAT', 'TAX ID', etc.
    prefixes_to_remove = ['CUI:', 'CUI', 'CIF:', 'CIF', 'VAT:', 'VAT', 'TAX ID:', 'TAX ID', 'NR.', 'NR', 'NO.', 'NO']
    for prefix in prefixes_to_remove:
        if vat.startswith(prefix):
            vat = vat[len(prefix):].strip()

    # Remove all spaces, dashes, dots, and other separators
    vat = re.sub(r'[\s\-\./:]+', '', vat)

    return vat


def extract_vat_numbers(vat: str) -> str:
    """Extract just the numeric portion of a VAT number for matching."""
    import re
    if not vat:
        return ''
    return re.sub(r'[^0-9]', '', str(vat))


def get_companies_with_vat() -> list[dict]:
    """Load companies with VAT numbers from the Comp sheet in Template."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True)

    if 'Comp' not in wb.sheetnames:
        wb.close()
        return []

    ws = wb['Comp']
    companies = []

    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).lower() if h else f'col_{j}' for j, h in enumerate(row)]
        else:
            if any(cell is not None for cell in row):
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = cell

                company_data = {
                    'company': row_dict.get('company', ''),
                    'brands': row_dict.get('brands', ''),
                    'vat': row_dict.get('vat', '')
                }
                if company_data['company']:
                    companies.append(company_data)

    wb.close()
    return companies


def match_company_by_vat(invoice_vat: str) -> Optional[dict]:
    """
    Find company matching the given VAT number.
    Uses multiple matching strategies:
    1. Exact match after normalization (removes spaces, prefixes)
    2. Numeric-only match (compares just the numbers)
    """
    if not invoice_vat:
        return None

    normalized_invoice_vat = normalize_vat(invoice_vat)
    invoice_numbers_only = extract_vat_numbers(invoice_vat)

    companies = get_companies_with_vat()

    # First pass: exact normalized match
    for company in companies:
        company_vat = company.get('vat', '')
        if normalize_vat(company_vat) == normalized_invoice_vat:
            return company

    # Second pass: numeric-only match (handles cases like 'RO225615' matching '225615')
    if invoice_numbers_only:
        for company in companies:
            company_vat = company.get('vat', '')
            company_numbers = extract_vat_numbers(company_vat)
            if company_numbers and company_numbers == invoice_numbers_only:
                return company

    return None


def add_company_with_vat(company: str, vat: str, brands: str = '') -> bool:
    """Add a new company with VAT to the Comp sheet."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    if 'Comp' not in wb.sheetnames:
        ws = wb.create_sheet('Comp')
        ws.cell(row=1, column=1, value='Company')
        ws.cell(row=1, column=2, value='Brands')
        ws.cell(row=1, column=3, value='VAT')
        next_row = 2
    else:
        ws = wb['Comp']
        next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1, value=company)
    ws.cell(row=next_row, column=2, value=brands)
    ws.cell(row=next_row, column=3, value=vat)

    wb.save(TEMPLATE_PATH)
    wb.close()
    return True


def update_company_vat(company: str, vat: str, brands: str = None) -> bool:
    """Update VAT for an existing company."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    if 'Comp' not in wb.sheetnames:
        wb.close()
        return False

    ws = wb['Comp']

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == company:
            ws.cell(row=row, column=3, value=vat)
            if brands is not None:
                ws.cell(row=row, column=2, value=brands)
            wb.save(TEMPLATE_PATH)
            wb.close()
            return True

    wb.close()
    return False


def delete_company(company: str) -> bool:
    """Delete a company from the Comp sheet."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    if 'Comp' not in wb.sheetnames:
        wb.close()
        return False

    ws = wb['Comp']

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == company:
            ws.delete_rows(row)
            wb.save(TEMPLATE_PATH)
            wb.close()
            return True

    wb.close()
    return False
